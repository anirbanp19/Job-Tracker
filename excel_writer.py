"""Excel output with two sheets: Strong Matches (>= threshold) and Other Matches."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def compute_combined_score(ats: int, fit: int) -> int:
    """Combined score: 60% ATS + 40% Fit. Used for ranking and threshold."""
    return round(0.6 * (ats or 0) + 0.4 * (fit or 0))


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
CELL_FONT = Font(name="Arial", size=10)
HYPERLINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Company", 22),
    ("Role", 32),
    ("Location", 18),
    ("Experience", 14),
    ("Min Years Required", 14),
    ("ATS Score", 10),
    ("Fit Score", 10),
    ("Combined Score", 12),
    ("Reasoning", 50),
    ("Missing Keywords", 40),
    ("Source(s)", 16),
    ("Apply Link", 12),
    ("City Priority", 12),
    ("Date Posted", 14),
    ("Date Found", 12),
    ("Date Last Seen", 14),
]


def _write_header(ws):
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, job: dict):
    sources = job.get("all_sources") or [(job.get("source", ""), job.get("url", ""))]
    primary_url = next((u for _, u in sources if u), "")
    source_str = ", ".join(sorted({(s or "").title() for s, _ in sources if s}))

    values = [
        job.get("company", ""),
        job.get("title", ""),
        job.get("location", ""),
        job.get("experience", ""),
        job.get("min_years_required") if job.get("min_years_required") is not None else "",
        job.get("ats_score", 0),
        job.get("fit_score", 0),
        compute_combined_score(job.get("ats_score", 0), job.get("fit_score", 0)),
        job.get("reasoning", ""),
        ", ".join(job.get("missing_keywords", [])),
        source_str,
        None,  # hyperlink filled below
        job.get("city_priority", ""),
        str(job.get("date_posted", ""))[:10],
        job.get("date_found", ""),
        job.get("date_last_seen", ""),
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.border = BORDER
        cell.alignment = CENTER_TOP if col_idx in (5, 9) else WRAP_TOP

    apply_col_idx = next(i for i, (name, _) in enumerate(COLUMNS, start=1) if name == "Apply Link")
    link_cell = ws.cell(row=row_idx, column=apply_col_idx, value="Apply")
    if primary_url:
        link_cell.hyperlink = primary_url
    link_cell.font = HYPERLINK_FONT
    link_cell.alignment = CENTER_TOP
    link_cell.border = BORDER


def _write_sheet(ws, title: str, jobs: list):
    ws.title = title
    _write_header(ws)
    # Sort: city priority asc, then score desc
    sorted_jobs = sorted(
        jobs,
        key=lambda j: (
            j.get("city_priority", 99),
            -compute_combined_score(j.get("ats_score", 0), j.get("fit_score", 0)),
        ),
    )
    for row_idx, job in enumerate(sorted_jobs, start=2):
        _write_row(ws, row_idx, job)


def write_excel(jobs: list, threshold: int, output_dir: str = "output", filename: str = "jobs_master.xlsx") -> str:
    os.makedirs(output_dir, exist_ok=True)

    strong = [j for j in jobs if compute_combined_score(j.get("ats_score", 0), j.get("fit_score", 0)) >= threshold]
    other = [j for j in jobs if compute_combined_score(j.get("ats_score", 0), j.get("fit_score", 0)) < threshold]

    wb = Workbook()
    _write_sheet(wb.active, "Strong Matches", strong)
    _write_sheet(wb.create_sheet("Other Matches"), "Other Matches", other)

    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    return out_path
